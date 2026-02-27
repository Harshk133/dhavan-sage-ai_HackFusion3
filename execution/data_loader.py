import os
import asyncio
from execution.db import get_database
import openpyxl

async def load_excel_to_mongo():
    db = await get_database()
    
    # Paths to excel files inside the provided /data directory
    med_path = "data/products-export.xlsx"  
    user_path = "data/Consumer Order History 1.xlsx"  

    print("Loading actual data into MongoDB using openpyxl...")

    # Load Medicines
    if os.path.exists(med_path):
        wb = openpyxl.load_workbook(med_path)
        ws = wb.active
        medicines_collection = db["medicines"]
        await medicines_collection.delete_many({})
        
        # Get Headers
        headers = [str(cell.value).lower() if cell.value else "" for cell in ws[1]]
        
        try:
            name_idx = headers.index("product name")
            pkg_idx = headers.index("package size")
        except ValueError:
            print("ERROR: Could not find required headers in products. Skipping.")
            name_idx, pkg_idx = None, None

        if name_idx is not None:
            parsed_meds = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[name_idx]: # only if product name exists
                    parsed_meds.append({
                        "name": str(row[name_idx]),
                        "dosage": str(row[pkg_idx]) if row[pkg_idx] else "Unknown",
                        "stock_level": 100, 
                        "prescription_required": False
                    })
                
            if parsed_meds:
                await medicines_collection.insert_many(parsed_meds)
                print(f"Loaded {len(parsed_meds)} medicines from products-export.xlsx.")
    else:
        print(f"File {med_path} not found. Skipping medicine init.")

    # Load Users
    if os.path.exists(user_path):
        wb = openpyxl.load_workbook(user_path)
        ws = wb.active
        users_collection = db["users"]
        await users_collection.delete_many({})
        
        # Header is on row 3 for User History
        headers = [str(cell.value).lower() if cell.value else "" for cell in ws[3]]
        
        patient_idx = None
        for i, h in enumerate(headers):
            if "patient" in h or "name" in h:
                patient_idx = i
                break
                
        parsed_users = []
        if patient_idx is not None:
            unique_patients = set()
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row[patient_idx]:
                    unique_patients.add(str(row[patient_idx]))
            
            for p in unique_patients:
                parsed_users.append({
                    "name": p,
                    "has_valid_prescription": True,
                    "purchase_history": []
                })
        else:
            print("WARNING: Could not identify a Patient column. Loading fallback users.")
            parsed_users = [
                 {"name": "John Doe", "has_valid_prescription": True, "purchase_history": []},
                 {"name": "Jane Smith", "has_valid_prescription": False, "purchase_history": []}
            ]

        if parsed_users:
            await users_collection.insert_many(parsed_users)
            print(f"Loaded {len(parsed_users)} users from Consumer Order History.")
    else:
        print(f"File {user_path} not found. Skipping user init.")

    print("Data loading complete.")

if __name__ == "__main__":
    asyncio.run(load_excel_to_mongo())
